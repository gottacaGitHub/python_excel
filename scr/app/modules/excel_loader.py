import os
from pathlib import Path

from .abstract_loader import AbstractExcelLoader
from ..core.constants import CONFIG
from ..core.dataclasses import ExcelFileInfo
from ..core.exceptions import (
    FileLoadError, FileFormatError,
    FileError, EmptyFileError
)


class ExcelLoader(AbstractExcelLoader):
    """Загрузчик .xlsx и .xls"""

    def load_file(self, file_path: str, has_headers: bool = CONFIG.has_headers) -> ExcelFileInfo:
        """
        Загружает данные

        Args:
            file_path: Путь к файлу
            has_headers: Заголовки в первой строке

        Returns:
            ExcelFileInfo: Информация о файле
        """
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"Файл не найден: {file_path}")

            if not self.validate_file_ext(file_path, CONFIG.excel_ext):
                raise FileFormatError(
                    f"Неподдерживаемый формат файла. "
                )

            file_ext = Path(file_path).suffix.lower()

            if file_ext == '.xlsx':
                return self._load_xlsx(file_path, has_headers)
            elif file_ext == '.xls':
                return self._load_xls(file_path, has_headers)
            else:
                raise FileFormatError(f"Расширение файла: {file_ext}")

        except (FileNotFoundError, FileFormatError, FileError, EmptyFileError):
            raise
        except Exception as e:
            raise FileError(f"Ошибка при чтении файла: {str(e)}")

    def _load_xlsx(self, file_path: str, has_headers: bool) -> ExcelFileInfo:
        """Загружает .xlsx"""
        try:
            from openpyxl import load_workbook

            file = load_workbook(filename=file_path, data_only=True, read_only=True)
            sheet = file.active

            data = []
            headers = []

            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0 and has_headers:
                    headers = [self.extract_value(cell) for cell in row]
                else:
                    data.append([self.extract_value(cell) for cell in row])

            if not data and not headers:
                raise EmptyFileError("Файл пуст")

            file.close()

            return ExcelFileInfo(
                file_path=file_path,
                file_name=Path(file_path).name,
                sheet_name=sheet.title,
                headers=headers,
                data=data,
                count_row=sheet.max_row,
                count_column=sheet.max_column
            )

        except ImportError:
            raise FileLoadError("Ошибка openpyxl")
        except Exception as e:
            raise FileError(f"Ошибка .xlsx файла: {str(e)}")

    def _load_xls(self, file_path: str, has_headers: bool) -> ExcelFileInfo:
        """Загружает .xls"""
        try:
            import xlrd

            file = xlrd.open_workbook(file_path, on_demand=True)
            sheet = file.sheet_by_index(0)

            data = []
            headers = []

            for i in range(sheet.nrows):
                row = sheet.row_values(i)
                if i == 0 and has_headers:
                    headers = [self.extract_value(cell) for cell in row]
                else:
                    data.append([self.extract_value(cell) for cell in row])

            if not data and not headers:
                raise EmptyFileError("Файл пуст")

            file.release_resources()

            return ExcelFileInfo(
                file_path=file_path,
                file_name=Path(file_path).name,
                sheet_name=sheet.name,
                headers=headers,
                data=data,
                count_row=sheet.nrows,
                count_column=sheet.ncols
            )

        except ImportError:
            raise FileLoadError("Ошибка xlrd")
        except Exception as e:
            raise FileError(f"Ошибка .xls: {str(e)}")